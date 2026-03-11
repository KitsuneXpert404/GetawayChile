from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse, HttpResponseForbidden
from django.db.models import Sum, Count, Q
from django.utils import timezone
import datetime
import json

from sales.models import Sale, PaymentStatus, Currency


def home(request):
    """Landing Page"""
    return render(request, 'core/home.html')


@login_required
def dashboard(request):
    """Role-based Dashboard Redirect & Data Loading"""
    user = request.user

    if user.is_admin:
        return render(request, 'core/dashboard_admin.html', _admin_context())

    if user.role == 'VENDEDOR':
        return render(request, 'core/dashboard_vendedor.html', _vendedor_context(user))

    if user.role == 'LOGISTICA':
        return redirect('logistics_dashboard')

    # Fallback
    return render(request, 'core/dashboard_admin.html', _admin_context())

def _vendedor_context(user):
    """Compute all KPIs and data for the salesperson dashboard (only their own data)."""
    today = timezone.now().date()
    first_of_month = today.replace(day=1)

    # --- KPIs ---
    my_sales_this_month = Sale.objects.filter(seller=user, created_at__date__gte=first_of_month)
    count_this_month = my_sales_this_month.count()
    revenue_clp = my_sales_this_month.filter(currency='CLP').aggregate(t=Sum('total_amount'))['t'] or 0
    revenue_usd = my_sales_this_month.filter(currency='USD').aggregate(t=Sum('total_amount'))['t'] or 0
    pax_this_month = my_sales_this_month.aggregate(t=Sum('passengers_count'))['t'] or 0

    # Growth vs previous month
    first_of_last_month = (first_of_month - datetime.timedelta(days=1)).replace(day=1)
    last_month_count = Sale.objects.filter(
        seller=user,
        created_at__date__gte=first_of_last_month,
        created_at__date__lt=first_of_month
    ).count()
    
    growth_pct = 0
    if last_month_count > 0:
        growth_pct = round(((count_this_month - last_month_count) / last_month_count) * 100, 1)

    # Today's pax for this seller
    today_pax = Sale.objects.filter(seller=user, tour_date=today).aggregate(t=Sum('passengers_count'))['t'] or 0

    # --- Payment status breakdown ---
    paid = my_sales_this_month.filter(payment_status='PAGADO').count()
    partial = my_sales_this_month.filter(payment_status='ABONADO').count()
    pending_pay = my_sales_this_month.filter(payment_status='PENDIENTE').count()

    # --- Per-day chart for current month (bar chart) ---
    chart_labels = []
    chart_counts = []
    chart_revenue = []

    day = first_of_month
    while day <= today:
        day_sales = Sale.objects.filter(seller=user, created_at__date=day)
        chart_labels.append(f"{day.day} {MESES_ES[day.month][:3]}")
        chart_counts.append(day_sales.count())
        rev = day_sales.filter(currency='CLP').aggregate(t=Sum('total_amount'))['t'] or 0
        chart_revenue.append(float(rev))
        day += datetime.timedelta(days=1)

    # --- Recent 5 sales ---
    recent_sales = Sale.objects.filter(seller=user).select_related('tour').order_by('-created_at')[:5]

    def fmt(n):
        try:
            return f"{int(n):,}".replace(',', '.')
        except Exception:
            return str(n)

    return {
        'kpis': {
            'count_this_month': count_this_month,
            'revenue_clp_fmt': fmt(int(revenue_clp)),
            'revenue_usd_fmt': fmt(int(revenue_usd)),
            'revenue_clp': float(revenue_clp),
            'revenue_usd': float(revenue_usd),
            'pax_this_month': pax_this_month,
            'growth_pct': growth_pct,
            'today_pax': today_pax,
            'paid': paid,
            'partial': partial,
            'pending_pay': pending_pay,
        },
        'chart_labels': json.dumps(chart_labels),
        'chart_counts': json.dumps(chart_counts),
        'chart_revenue': json.dumps(chart_revenue),
        'month_name': f"{MESES_ES[today.month]} {today.year}",
        'updated_time': timezone.localtime(timezone.now()).strftime('%H:%M'),
        'recent_sales': recent_sales,
    }


MESES_ES = {
    1:'Enero',2:'Febrero',3:'Marzo',4:'Abril',5:'Mayo',6:'Junio',
    7:'Julio',8:'Agosto',9:'Septiembre',10:'Octubre',11:'Noviembre',12:'Diciembre'
}
DIAS_ES = {0:'Lun',1:'Mar',2:'Mié',3:'Jue',4:'Vie',5:'Sáb',6:'Dom'}


def _admin_context():
    """Compute all KPIs and data for the admin dashboard."""
    today = timezone.now().date()
    first_of_month = today.replace(day=1)

    # --- KPIs ---
    sales_this_month = Sale.objects.filter(created_at__date__gte=first_of_month)
    count_this_month = sales_this_month.count()
    revenue_clp = sales_this_month.filter(currency='CLP').aggregate(t=Sum('total_amount'))['t'] or 0
    revenue_usd = sales_this_month.filter(currency='USD').aggregate(t=Sum('total_amount'))['t'] or 0
    pax_this_month = sales_this_month.aggregate(t=Sum('passengers_count'))['t'] or 0

    # Growth vs previous month
    first_of_last_month = (first_of_month - datetime.timedelta(days=1)).replace(day=1)
    last_month_count = Sale.objects.filter(
        created_at__date__gte=first_of_last_month,
        created_at__date__lt=first_of_month
    ).count()
    growth_pct = 0
    if last_month_count > 0:
        growth_pct = round(((count_this_month - last_month_count) / last_month_count) * 100, 1)

    # Pending tickets
    try:
        from tickets.models import Ticket, TicketStatus
        pending_tickets = Ticket.objects.filter(status='PENDIENTE').count()
    except Exception:
        pending_tickets = 0

    # Today's tours
    today_tours = Sale.objects.filter(tour_date=today).count()

    # --- Payment status breakdown ---
    paid = sales_this_month.filter(payment_status='PAGADO').count()
    partial = sales_this_month.filter(payment_status='ABONADO').count()
    pending_pay = sales_this_month.filter(payment_status='PENDIENTE').count()

    # --- Top sellers ---
    # We rank by number of sales to avoid mixing currencies (CLP, USD, BRL) in a sum.
    from django.db.models import Q
    top_sellers = (
        Sale.objects
        .filter(created_at__date__gte=first_of_month, seller__isnull=False)
        .values('seller__id', 'seller__first_name', 'seller__last_name')
        .annotate(
            total_clp=Sum('total_amount', filter=Q(currency='CLP')),
            num_sales=Count('id')
        )
        .order_by('-num_sales')[:5]
    )
    max_sales = max((s['num_sales'] for s in top_sellers), default=1)
    sellers_ranking = [
        {
            'name': f"{s['seller__first_name']} {s['seller__last_name']}".strip(),
            'total_clp': s['total_clp'] or 0,
            'num_sales': s['num_sales'],
            'pct': round((s['num_sales'] / max_sales) * 100) if max_sales else 0,
        }
        for s in top_sellers
    ]

    # --- Per-day chart for current month (bar chart) ---
    chart_labels = []
    chart_counts = []
    chart_revenue = []

    # Generate one point per day in the current month up to today
    day = first_of_month
    while day <= today:
        day_sales = Sale.objects.filter(created_at__date=day)
        chart_labels.append(f"{day.day} {MESES_ES[day.month][:3]}")
        chart_counts.append(day_sales.count())
        rev = day_sales.filter(currency='CLP').aggregate(t=Sum('total_amount'))['t'] or 0
        chart_revenue.append(float(rev))
        day += datetime.timedelta(days=1)

    # --- Recent 5 sales ---
    recent_sales = Sale.objects.select_related('tour', 'seller').order_by('-created_at')[:5]

    def fmt(n):
        """Format number as Chilean peso string without requiring intcomma filter."""
        try:
            return f"{int(n):,}".replace(',', '.')
        except Exception:
            return str(n)

    for s in sellers_ranking:
        s['total_fmt'] = fmt(s.get('total_clp', 0))

    return {
        'kpis': {
            'count_this_month': count_this_month,
            'revenue_clp_fmt': fmt(int(revenue_clp)),
            'revenue_usd_fmt': fmt(int(revenue_usd)),
            'revenue_clp': float(revenue_clp),
            'revenue_usd': float(revenue_usd),
            'pax_this_month': pax_this_month,
            'growth_pct': growth_pct,
            'pending_tickets': pending_tickets,
            'today_tours': today_tours,
            'paid': paid,
            'partial': partial,
            'pending_pay': pending_pay,
        },
        'sellers_ranking': sellers_ranking,
        'chart_labels': json.dumps(chart_labels),
        'chart_counts': json.dumps(chart_counts),
        'chart_revenue': json.dumps(chart_revenue),
        'month_name': f"{MESES_ES[today.month]} {today.year}",
        'updated_time': timezone.localtime(timezone.now()).strftime('%H:%M'),
        'recent_sales': recent_sales,
    }



# --- Report Exports ---

def _user_can_export_sales(user):
    """Solo ADMIN y LOGISTICA pueden exportar ventas a Excel."""
    if not user or not user.is_authenticated:
        return False
    return getattr(user, 'is_admin', False) or user.role in ('ADMIN', 'LOGISTICA')


@login_required
def export_sales_excel(request):
    """Export sales of the current month as an Excel file. Solo ADMIN y LOGISTICA."""
    if not _user_can_export_sales(request.user):
        return HttpResponseForbidden("No tiene permiso para exportar ventas.")
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return HttpResponse("openpyxl no está instalado. Ejecuta: pip install openpyxl", status=500)

    today = timezone.now().date()
    first_of_month = today.replace(day=1)
    month_str = today.strftime('%B_%Y')

    sales = Sale.objects.filter(
        created_at__date__gte=first_of_month
    ).select_related('tour', 'seller').order_by('-created_at')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Ventas {today.strftime('%B %Y')}"

    # Header
    headers = ['#ID', 'Fecha', 'Vendedor', 'Cliente', 'Tour', 'Fecha Tour', 'Pax', 'Total', 'Moneda', 'Estado Pago']
    header_fill = PatternFill(start_color='1a56db', end_color='1a56db', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    # Rows
    for row, sale in enumerate(sales, 2):
        ws.cell(row=row, column=1, value=sale.id)
        ws.cell(row=row, column=2, value=sale.created_at.strftime('%d/%m/%Y %H:%M'))
        seller = f"{sale.seller.first_name} {sale.seller.last_name}".strip() if sale.seller else 'Web'
        ws.cell(row=row, column=3, value=seller)
        ws.cell(row=row, column=4, value=f"{sale.client_first_name} {sale.client_last_name}")
        ws.cell(row=row, column=5, value=sale.tour.name if sale.tour else '-')
        ws.cell(row=row, column=6, value=sale.tour_date.strftime('%d/%m/%Y') if sale.tour_date else '-')
        ws.cell(row=row, column=7, value=sale.passengers_count)
        ws.cell(row=row, column=8, value=float(sale.total_amount))
        ws.cell(row=row, column=9, value=sale.currency)
        ws.cell(row=row, column=10, value=sale.payment_status)

    # Auto-width
    for col in ws.columns:
        max_len = max((len(str(c.value or '')) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="Ventas_Getaway_{month_str}.xlsx"'
    wb.save(response)
    return response
